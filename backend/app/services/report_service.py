"""
ReportService – generates CSV, XLSX, and TXT reports from persisted events.
Saves files to the configured REPORTS_DIR.
"""
import csv
import os
from datetime import datetime
from pathlib import Path
from typing import Literal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.config import get_settings
from app.core.logging import get_logger
from app.models.attack_event import AttackEvent
from app.schemas.stats import StatsResponse

logger = get_logger(__name__)

ReportFormat = Literal["csv", "xlsx", "txt"]


class ReportService:
    def __init__(self):
        self._dir = Path(get_settings().REPORTS_DIR)
        self._dir.mkdir(parents=True, exist_ok=True)

    # ── Public ────────────────────────────────────────────────────────────────

    def generate(
        self,
        events: list[AttackEvent],
        stats: StatsResponse,
        fmt: ReportFormat = "csv",
    ) -> Path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = self._dir / f"attack_report_{ts}.{fmt}"
        generators = {"csv": self._csv, "xlsx": self._xlsx, "txt": self._txt}
        generators[fmt](events, stats, path)
        logger.info("Report generated: %s (%d events)", path.name, len(events))
        return path

    def safe_path(self, filename: str) -> Path:
        """Resolve and validate that a filename stays inside REPORTS_DIR."""
        resolved = (self._dir / filename).resolve()
        if not str(resolved).startswith(str(self._dir.resolve())):
            raise ValueError("Path traversal detected.")
        return resolved

    # ── Private generators ────────────────────────────────────────────────────

    @staticmethod
    def _csv(events: list[AttackEvent], _stats: StatsResponse, path: Path) -> None:
        FIELDS = [
            "id", "timestamp", "service", "source_ip", "source_port",
            "username", "severity", "ai_label", "threat_score", "command",
        ]
        with path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=FIELDS, extrasaction="ignore")
            writer.writeheader()
            for e in events:
                writer.writerow({f: getattr(e, f, "") for f in FIELDS})

    @staticmethod
    def _txt(events: list[AttackEvent], stats: StatsResponse, path: Path) -> None:
        SEP = "=" * 70
        with path.open("w", encoding="utf-8") as fh:
            fh.write(f"{SEP}\nHoneyCloud Security Report\n")
            fh.write(f"Generated : {datetime.now():%Y-%m-%d %H:%M:%S}\n{SEP}\n\n")

            fh.write("EXECUTIVE SUMMARY\n" + "-" * 70 + "\n")
            fh.write(f"Total Events : {stats.total_events}\n\n")

            fh.write("By Service:\n")
            for svc, cnt in stats.events_by_service.items():
                fh.write(f"  {svc:10s}: {cnt}\n")

            fh.write("\nBy Severity:\n")
            for sev, cnt in stats.events_by_severity.items():
                fh.write(f"  {sev:10s}: {cnt}\n")

            fh.write("\nAI Labels:\n")
            for lbl, cnt in stats.ai_labels.items():
                fh.write(f"  {lbl:10s}: {cnt}\n")

            fh.write(f"\n{SEP}\nRECENT EVENTS (last 20)\n{SEP}\n\n")
            for e in events[:20]:
                geo = e.geolocation or {}
                fh.write(
                    f"[{e.id}] {e.timestamp} | {e.service} | {e.source_ip} "
                    f"| {geo.get('city','?')}, {geo.get('country','?')}\n"
                    f"     Severity={e.severity}  Label={e.ai_label}  "
                    f"Score={e.threat_score:.2f}  User={e.username or 'N/A'}\n"
                )
                if e.command:
                    fh.write(f"     Command: {e.command}\n")
                fh.write("\n")

    @staticmethod
    def _xlsx(events: list[AttackEvent], stats: StatsResponse, path: Path) -> None:
        wb = openpyxl.Workbook()

        # ── Summary sheet ─────────────────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum.column_dimensions["A"].width = 22
        ws_sum.column_dimensions["B"].width = 16

        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(fill_type="solid", fgColor="0A0E27")

        ws_sum["A1"] = "HoneyCloud Attack Report"
        ws_sum["A1"].font = header_font
        ws_sum["A1"].fill = header_fill
        ws_sum.merge_cells("A1:B1")

        ws_sum["A2"] = f"Generated: {datetime.now():%Y-%m-%d %H:%M:%S}"
        ws_sum["A2"].font = Font(italic=True, size=10)
        ws_sum.merge_cells("A2:B2")

        rows = [
            ("Total Events", stats.total_events),
            ("Critical", stats.events_by_severity.get("CRITICAL", 0)),
            ("High",     stats.events_by_severity.get("HIGH", 0)),
            ("Malicious (AI)", stats.ai_labels.get("malicious", 0)),
        ]
        for r, (label, val) in enumerate(rows, start=4):
            ws_sum.cell(r, 1, label).font = Font(bold=True)
            ws_sum.cell(r, 2, val)

        # ── Events sheet ──────────────────────────────────────────────────────
        ws_ev = wb.create_sheet("Events")
        HEADERS = ["ID", "Timestamp", "Service", "Source IP", "Username",
                   "Severity", "AI Label", "Score", "Command", "City", "Country"]
        COL_W   = [8, 22, 10, 16, 14, 10, 12, 8, 30, 14, 14]

        hdr_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        hdr_font = Font(bold=True, color="FFFFFF")

        for col, (hdr, w) in enumerate(zip(HEADERS, COL_W), start=1):
            cell = ws_ev.cell(1, col, hdr)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            ws_ev.column_dimensions[chr(64 + col)].width = w

        SEV_COLORS = {"CRITICAL": "FFCCCC", "HIGH": "FFE5CC", "MEDIUM": "FFFACC"}
        for row, e in enumerate(events, start=2):
            geo = e.geolocation or {}
            values = [
                e.id, str(e.timestamp), e.service, e.source_ip,
                e.username, e.severity, e.ai_label, e.threat_score,
                e.command, geo.get("city"), geo.get("country"),
            ]
            for col, val in enumerate(values, start=1):
                ws_ev.cell(row, col, val)
            if e.severity in SEV_COLORS:
                fill = PatternFill(fill_type="solid", fgColor=SEV_COLORS[e.severity])
                ws_ev.cell(row, 6).fill = fill

        wb.save(path)
